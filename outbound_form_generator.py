"""
出库单生成器
相同出库单号、出库日期、仓库类型的物料可以合并到一张出库单
每个物料可能有多次出库（批次1、批次2、批次3）
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import os


class OutboundFormGenerator:
    """出库单生成器"""

    def __init__(self, data_file: str, output_dir: str, template_file: str = None):
        """
        初始化生成器

        Args:
            data_file: 数据明细表文件路径
            output_dir: 输出目录
            template_file: 出库单模板文件路径
        """
        self.data_file = data_file
        self.output_dir = output_dir
        self.template_file = template_file or "assets/出库单单模板.xlsx"

        # 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)

        # 加载模板
        self.template_wb = load_workbook(self.template_file)
        self.template_ws = self.template_wb.active

        # 加载数据
        self.df = pd.read_excel(data_file, header=3)
        self.df = self.df.dropna(how='all')

        # 建立列名到列索引的映射
        self.column_map = {}
        for idx, col_name in enumerate(self.df.columns):
            self.column_map[str(col_name).strip()] = idx

    def _get_column_index(self, possible_names, fallback_idx):
        """根据可能的列名查找列索引"""
        for name in possible_names:
            for col_name in self.df.columns:
                if name in str(col_name):
                    return self.df.columns.get_loc(col_name)
        return fallback_idx if fallback_idx < len(self.df.columns) else 0

    def _safe_float(self, value):
        """安全地将值转换为float"""
        if pd.isna(value):
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    def generate_all(self):
        """
        生成所有出库单（按出库单号、出库日期、仓库类型分组）

        Returns:
            生成结果字典
        """
        print(f"开始生成出库单，共 {len(self.df)} 个物料")

        result = {
            'total': 0,
            'outbound_count': 0,
            'success': 0,
            'failed': 0,
            'forms': [],
            'errors': []
        }

        # 收集所有出库记录（包括多个批次）
        outbound_records = self._collect_outbound_records()
        result['outbound_count'] = len(outbound_records)

        print(f"共收集 {len(outbound_records)} 条出库记录")

        # 按出库单号、出库日期、仓库类型分组
        groups = self._group_outbound_records(outbound_records)

        print(f"共生成 {len(groups)} 张出库单")

        for group_key, group_data in groups.items():
            try:
                file_path = self._generate_one(group_data, group_key)
                result['total'] += 1
                result['success'] += 1
                result['forms'].append({
                    'outbound_no': group_key[0],
                    'outbound_date': group_key[1],
                    'warehouse': group_key[2],
                    'material_count': len(group_data),
                    'file_path': file_path
                })
                print(f"  ✓ 生成成功: 出库单号 {group_key[0]}, 仓库 {group_key[2]}, 物料数 {len(group_data)}")
            except Exception as e:
                result['failed'] += 1
                import traceback
                error_details = traceback.format_exc()
                error_msg = f"出库单 {group_key[0]} 生成失败: {str(e)}\n详细堆栈:\n{error_details}"
                result['errors'].append(error_msg)
                print(f"  ✗ 出库单 {group_key[0]} 生成失败: {str(e)}")

        print(f"\n出库单生成完成！成功: {result['success']}, 失败: {result['failed']}")
        return result

    def _collect_outbound_records(self):
        """
        收集所有出库记录（每个物料的多个批次展开为多条记录）

        Returns:
            出库记录列表
        """
        records = []

        # 列索引（根据数据明细表测试.xlsx的列结构）
        material_code_col = 1  # 物料编码
        material_name_col = 2  # 物料名称
        brand_col = 5  # 品牌
        spec_col = 3  # 规格
        unit_col = 4  # 单位
        batch_col = 19  # 入库批号
        expiry_col = 21  # 有效期
        warehouse_col = 22  # 仓库

        # 批次1
        out_date1_col = 24  # 领料&出库日期
        out_no1_col = 25  # 出库单号
        out_qty1_col = 26  # 领料量

        # 批次2
        out_date2_col = 28  # 领料&出库日期.1
        out_no2_col = 29  # 出库单号.1
        out_qty2_col = 30  # 领料量.1

        # 批次3
        out_date3_col = 32  # 领料&出库日期.2
        out_no3_col = 33  # 出库单号.2
        out_qty3_col = 34  # 领料量.2

        print(f"    使用列索引: 批次1(日期={out_date1_col}, 单号={out_no1_col}, 数量={out_qty1_col})")

        for idx, row in self.df.iterrows():
            # 获取物料基本信息
            material_code = self._get_value(row, material_code_col)
            material_name = self._get_value(row, material_name_col)
            brand = self._get_value(row, brand_col)
            specification = self._get_value(row, spec_col)
            unit = self._get_value(row, unit_col)
            batch_no = self._get_value(row, batch_col)
            expiry_date = self._get_value(row, expiry_col)
            warehouse = self._get_value(row, warehouse_col)

            # 批次1
            out_date1 = self._get_value(row, out_date1_col)
            out_no1 = self._get_value(row, out_no1_col)
            out_qty1 = self._safe_float(self._get_value(row, out_qty1_col))
            if out_no1 and out_qty1 > 0:
                records.append({
                    'material_code': material_code,
                    'material_name': material_name,
                    'brand': brand,
                    'specification': specification,
                    'unit': unit,
                    'batch_no': batch_no,
                    'expiry_date': expiry_date,
                    'warehouse': warehouse,
                    'outbound_no': out_no1,
                    'outbound_date': out_date1,
                    'quantity': out_qty1,
                    'remark': ''
                })

            # 批次2
            out_date2 = self._get_value(row, out_date2_col)
            out_no2 = self._get_value(row, out_no2_col)
            out_qty2 = self._safe_float(self._get_value(row, out_qty2_col))
            if out_no2 and out_qty2 > 0:
                records.append({
                    'material_code': material_code,
                    'material_name': material_name,
                    'brand': brand,
                    'specification': specification,
                    'unit': unit,
                    'batch_no': batch_no,
                    'expiry_date': expiry_date,
                    'warehouse': warehouse,
                    'outbound_no': out_no2,
                    'outbound_date': out_date2,
                    'quantity': out_qty2,
                    'remark': ''
                })

            # 批次3
            out_date3 = self._get_value(row, out_date3_col)
            out_no3 = self._get_value(row, out_no3_col)
            out_qty3 = self._safe_float(self._get_value(row, out_qty3_col))
            if out_no3 and out_qty3 > 0:
                records.append({
                    'material_code': material_code,
                    'material_name': material_name,
                    'brand': brand,
                    'specification': specification,
                    'unit': unit,
                    'batch_no': batch_no,
                    'expiry_date': expiry_date,
                    'warehouse': warehouse,
                    'outbound_no': out_no3,
                    'outbound_date': out_date3,
                    'quantity': out_qty3,
                    'remark': ''
                })

        return records

    def _group_outbound_records(self, records):
        """
        按出库单号、出库日期、仓库类型分组

        Returns:
            分组字典 {(出库单号, 出库日期, 仓库类型): [记录列表]}
        """
        groups = {}

        for record in records:
            # 获取分组键
            outbound_no = record['outbound_no']
            outbound_date = self._format_date(record['outbound_date'])
            warehouse = record['warehouse']

            # 跳过没有出库单号的记录
            if not outbound_no:
                continue

            group_key = (outbound_no, outbound_date, warehouse)

            if group_key not in groups:
                groups[group_key] = []
            groups[group_key].append(record)

        return groups

    def _generate_one(self, group_data, group_key):
        """
        生成一张出库单

        Args:
            group_data: 分组数据列表
            group_key: 分组键 (出库单号, 出库日期, 仓库类型)

        Returns:
            生成的文件路径
        """
        # 复制模板
        wb = load_workbook(self.template_file)
        ws = wb.active

        outbound_no, outbound_date, warehouse = group_key

        # 填充表头信息（根据模板文件的结构）
        # 行3: 公司名称(C3)、出库类型(E3)、出库单号(H3)
        ws['C3'] = '深圳市海普洛斯生物科技有限公司'
        ws['E3'] = 'IVD材料出库'
        ws['H3'] = outbound_no

        # 行4: 仓库(C4)、领用部门(E4)、出库日期(H4)
        ws['C4'] = warehouse
        ws['E4'] = '生产部'
        ws['H4'] = outbound_date

        # 填充物料明细（从第6行开始，表头在第5行）
        start_row = 6
        for i, record in enumerate(group_data):
            row_idx = start_row + i

            # 序号：10, 20, 30...
            seq = (i + 1) * 10

            # 填充数据（根据模板文件的列结构）
            ws.cell(row=row_idx, column=1, value=seq)  # 序号（A列）
            ws.cell(row=row_idx, column=2, value=record['material_code'])  # 物料编码（B列）
            ws.cell(row=row_idx, column=3, value=record['material_name'])  # 物料名称（C列）
            ws.cell(row=row_idx, column=4, value=record['brand'])  # 品牌（D列）
            ws.cell(row=row_idx, column=5, value=record['specification'])  # 规格（E列）
            ws.cell(row=row_idx, column=6, value=record['unit'])  # 单位（F列）
            ws.cell(row=row_idx, column=7, value=self._format_number(record['quantity']))  # 数量（G列）
            ws.cell(row=row_idx, column=8, value=record['batch_no'])  # 批次（H列）
            ws.cell(row=row_idx, column=9, value=self._format_date(record['expiry_date']))  # 失效日期（I列）
            ws.cell(row=row_idx, column=10, value=record['remark'])  # 备注（J列）

        # 生成文件名
        filename = f"出库单_{outbound_no}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(self.output_dir, filename)

        # 保存文件
        wb.save(file_path)

        return file_path

    def _get_value(self, row, col_idx):
        """获取单元格值"""
        if col_idx >= len(row):
            return None
        val = row.iloc[col_idx]
        if pd.isna(val):
            return None
        return val

    def _format_date(self, date_value):
        """格式化日期"""
        if pd.isna(date_value):
            return ''
        if isinstance(date_value, datetime):
            return date_value.strftime('%Y-%m-%d')
        if isinstance(date_value, str):
            # 尝试解析
            for fmt in ['%Y.%m.%d', '%Y-%m-%d', '%Y/%m/%d']:
                try:
                    dt = datetime.strptime(date_value, fmt)
                    return dt.strftime('%Y-%m-%d')
                except:
                    continue
        return str(date_value)

    def _format_number(self, num_value):
        """格式化数字，保留两位小数"""
        if pd.isna(num_value):
            return ''
        if isinstance(num_value, (int, float)):
            return f"{float(num_value):.2f}"
        if isinstance(num_value, str) and num_value.strip():
            # 尝试转换
            try:
                return f"{float(num_value):.2f}"
            except:
                return num_value
        return ''
