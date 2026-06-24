"""
请验单生成器
每个物料对应一个单独的请验单
"""

import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os


class InspectionFormGenerator:
    """请验单生成器"""

    def __init__(self, data_file: str, output_dir: str, template_file: str = None, header_row: int = 3):
        """
        初始化生成器

        Args:
            data_file: 数据明细表文件路径
            output_dir: 输出目录
            template_file: 请验单模板文件路径
            header_row: 表头所在的行索引（默认为3，即第4行）
        """
        self.data_file = data_file
        self.output_dir = output_dir
        self.template_file = template_file or "assets/请验单模板.xlsx"
        self.header_row = header_row

        # 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)

        # 加载数据
        self.df = pd.read_excel(data_file, header=header_row)
        self.df = self.df.dropna(how='all')

    def generate_all(self):
        """
        为所有物料生成请验单

        Returns:
            生成结果字典
        """
        print(f"开始生成请验单，共 {len(self.df)} 个物料")

        result = {
            'total': len(self.df),
            'success': 0,
            'failed': 0,
            'files': [],
            'errors': []
        }

        for idx, row in self.df.iterrows():
            try:
                file_path = self._generate_one(row, idx)
                result['success'] += 1
                result['files'].append({
                    'material_code': self._get_value(row, 1),  # 物料编码
                    'material_name': self._get_value(row, 2),  # 物料名称
                    'file_path': file_path
                })
                print(f"  ✓ 生成成功: {self._get_value(row, 1)} - {self._get_value(row, 2)}")
            except Exception as e:
                result['failed'] += 1
                error_msg = f"物料 {int(idx)+1} 生成失败: {str(e)}"
                result['errors'].append(error_msg)
                print(f"  ✗ {error_msg}")

        print(f"\n请验单生成完成！成功: {result['success']}, 失败: {result['failed']}")
        return result

    def _generate_one(self, row, idx):
        """
        为单个物料生成请验单

        Args:
            row: 数据行
            idx: 行索引

        Returns:
            生成的文件路径
        """
        # 复制模板
        wb = load_workbook(self.template_file)
        ws = wb.active

        # 列索引（根据数据明细表测试.xlsx的列结构）
        material_code_col = 1  # 物料编码
        material_name_col = 2  # 物料名称
        sample_category_col = 7  # 样品分类
        batch_no_col = 19  # 入库批号
        inspection_date_col = 15  # 到货&请验单日期
        specification_col = 3  # 规格
        inbound_qty_col = 8  # 采购数量
        sample_qty_col = 13  # 抽样数量
        unit_col = 4  # 单位

        # 获取物料数据
        material_code = self._get_value(row, material_code_col)
        material_name = self._get_value(row, material_name_col)
        sample_category = self._get_value(row, sample_category_col)
        batch_no = self._get_value(row, batch_no_col)
        inspection_date = self._get_value(row, inspection_date_col)
        specification = self._get_value(row, specification_col)
        inbound_qty = self._get_value(row, inbound_qty_col)
        sample_qty = self._get_value(row, sample_qty_col)
        unit = self._get_value(row, unit_col)

        # 构建要填充的数据
        fill_data = {
            'material_name': material_name,
            'sample_category': sample_category,
            'batch_no': str(batch_no) if batch_no else '',
            'inspection_department': '仓储部',
            'inspector': '小明',
            'inspection_date': self._format_date(inspection_date),
            'specification': specification,
            'quantity': f"{self._format_number(inbound_qty)}{unit}" if inbound_qty and unit else '',
            'sampler': '小红',
            'sample_qty': f"{self._format_number(sample_qty)}{unit}" if sample_qty and unit else '',
            'material_code': material_code,
            'remark': '抽样检验消耗'
        }

        # 填充上联（行4-9）
        self._fill_inspection_form(ws, fill_data, start_row=4)

        # 填充下联（行16-21）
        self._fill_inspection_form(ws, fill_data, start_row=16)

        # 生成文件名
        filename = f"请验单_{material_code}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(self.output_dir, filename)

        # 保存文件
        wb.save(file_path)

        return file_path

    def _fill_inspection_form(self, ws, data, start_row):
        """
        填充一个请验单（上联或下联）

        Args:
            ws: 工作表
            data: 要填充的数据字典
            start_row: 起始行号
        """
        # 行+0: 样品名称(B)、样品分类(D)
        ws.cell(row=start_row, column=2, value=data['material_name'])
        ws.cell(row=start_row, column=4, value=data['sample_category'])

        # 行+1: 批号(B)、请验部门(D)
        ws.cell(row=start_row + 1, column=2, value=data['batch_no'])
        ws.cell(row=start_row + 1, column=4, value=data['inspection_department'])

        # 行+2: 请验人(B)、请验日期(D)
        ws.cell(row=start_row + 2, column=2, value=data['inspector'])
        ws.cell(row=start_row + 2, column=4, value=data['inspection_date'])

        # 行+3: 规格(B)、数量(D)
        ws.cell(row=start_row + 3, column=2, value=data['specification'])
        ws.cell(row=start_row + 3, column=4, value=data['quantity'])

        # 行+4: 抽样人(B)、抽样数量(D)
        ws.cell(row=start_row + 4, column=2, value=data['sampler'])
        ws.cell(row=start_row + 4, column=4, value=data['sample_qty'])

        # 行+5: 物料编码(B)、备注(D)
        ws.cell(row=start_row + 5, column=2, value=data['material_code'])
        ws.cell(row=start_row + 5, column=4, value=data['remark'])

    def _get_value(self, row, col_idx):
        """获取单元格值"""
        if col_idx >= len(row):
            return None
        val = row.iloc[col_idx]
        if pd.isna(val):
            return None
        return val

    def _format_date(self, date_value):
        """格式化日期"""
        if pd.isna(date_value):
            return ''
        if isinstance(date_value, datetime):
            return date_value.strftime('%Y.%m.%d')
        if isinstance(date_value, str):
            for fmt in ['%Y.%m.%d', '%Y-%m-%d', '%Y/%m/%d']:
                try:
                    dt = datetime.strptime(date_value, fmt)
                    return dt.strftime('%Y.%m.%d')
                except:
                    continue
        return str(date_value)

    def _format_number(self, num_value):
        """格式化数字，保留两位小数"""
        if pd.isna(num_value):
            return ''
        if isinstance(num_value, (int, float)):
            return f"{float(num_value):.2f}"
        if isinstance(num_value, str) and num_value.strip():
            try:
                return f"{float(num_value):.2f}"
            except:
                return num_value
        return ''
