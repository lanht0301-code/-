"""
表单核验工具
用于验证生成的表单与数据明细表的数据是否一一对应
"""

import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os
import glob


class FormValidator:
    """表单核验器"""

    def __init__(self, data_file: str, output_dir: str):
        """
        初始化核验器

        Args:
            data_file: 数据明细表文件路径
            output_dir: 生成表单的输出目录
        """
        self.data_file = data_file
        self.output_dir = output_dir

        # 加载数据明细表
        self.df = pd.read_excel(data_file, header=3)
        self.df = self.df.dropna(how='all')

        self.validation_results = []

    def validate_all(self):
        """
        核验所有生成的表单

        Returns:
            核验结果字典
        """
        print(f"开始核验表单，数据明细表共 {len(self.df)} 行")

        result = {
            'inspection_forms': self._validate_inspection_forms(),
            'inbound_forms': self._validate_inbound_forms(),
            'outbound_forms': self._validate_outbound_forms(),
            'summary': {}
        }

        # 生成汇总
        result['summary'] = {
            'total_checks': (
                result['inspection_forms']['total'] +
                result['inbound_forms']['total'] +
                result['outbound_forms']['total']
            ),
            'total_passed': (
                result['inspection_forms']['passed'] +
                result['inbound_forms']['passed'] +
                result['outbound_forms']['passed']
            ),
            'total_failed': (
                result['inspection_forms']['failed'] +
                result['inbound_forms']['failed'] +
                result['outbound_forms']['failed']
            ),
            'total_warnings': (
                result['inspection_forms']['warnings'] +
                result['inbound_forms']['warnings'] +
                result['outbound_forms']['warnings']
            )
        }

        print(f"\n表单核验完成！")
        print(f"总检查项: {result['summary']['total_checks']}")
        print(f"通过: {result['summary']['total_passed']}")
        print(f"失败: {result['summary']['total_failed']}")
        print(f"警告: {result['summary']['total_warnings']}")

        return result

    def _validate_inspection_forms(self):
        """核验请验单"""
        print("\n  核验请验单...")

        result = {
            'total': 0,
            'passed': 0,
            'failed': 0,
            'warnings': 0,
            'details': []
        }

        # 查找所有请验单文件
        pattern = os.path.join(self.output_dir, "请验单_*.xlsx")
        files = glob.glob(pattern)

        if not files:
            result['warnings'] = len(self.df)
            result['details'].append({
                'type': 'warning',
                'message': f"未找到请验单文件（预期约 {len(self.df)} 个）"
            })
            return result

        # 核验每个请验单
        for file_path in files:
            result['total'] += 1
            try:
                details = self._validate_one_inspection_form(file_path)
                if details['status'] == 'passed':
                    result['passed'] += 1
                elif details['status'] == 'failed':
                    result['failed'] += 1
                result['details'].append(details)
            except Exception as e:
                result['failed'] += 1
                result['details'].append({
                    'type': 'failed',
                    'file': file_path,
                    'message': f"核验失败: {str(e)}"
                })

        print(f"    请验单核验完成: 总数 {result['total']}, 通过 {result['passed']}, 失败 {result['failed']}")
        return result

    def _validate_one_inspection_form(self, file_path):
        """核验单个请验单"""
        wb = load_workbook(file_path)
        ws = wb.active

        # 从文件名提取物料编码
        filename = os.path.basename(file_path)
        material_code = filename.split('_')[1] if '_' in filename else None

        if not material_code:
            return {
                'type': 'failed',
                'file': file_path,
                'message': f"无法从文件名提取物料编码"
            }

        # 从数据明细表查找对应物料
        material_rows = self.df[self.df.iloc[:, 1].astype(str) == str(material_code)]

        if len(material_rows) == 0:
            return {
                'type': 'failed',
                'file': file_path,
                'message': f"数据明细表中未找到物料编码 {material_code}"
            }

        row = material_rows.iloc[0]
        errors = []

        # 核验各个字段
        # 样品名称 (C列, 索引2)
        expected_name = self._get_value(row, 2)
        actual_name = ws['B4'].value
        if str(expected_name) != str(actual_name):
            errors.append(f"样品名称不匹配: 预期'{expected_name}', 实际'{actual_name}'")

        # 物料编码 (B列, 索引1)
        expected_code = self._get_value(row, 1)
        actual_code = ws['B9'].value
        if str(expected_code) != str(actual_code):
            errors.append(f"物料编码不匹配: 预期'{expected_code}', 实际'{actual_code}'")

        # 批号 (T列, 索引19)
        expected_batch = self._get_value(row, 19)
        actual_batch = ws['B5'].value
        if str(expected_batch or '') != str(actual_batch or ''):
            errors.append(f"批号不匹配: 预期'{expected_batch}', 实际'{actual_batch}'")

        if errors:
            return {
                'type': 'failed',
                'file': file_path,
                'material_code': material_code,
                'errors': errors
            }

        return {
            'type': 'passed',
            'file': file_path,
            'material_code': material_code
        }

    def _validate_inbound_forms(self):
        """核验入库单"""
        print("\n  核验入库单...")

        result = {
            'total': 0,
            'passed': 0,
            'failed': 0,
            'warnings': 0,
            'details': []
        }

        # 查找所有入库单文件
        pattern = os.path.join(self.output_dir, "入库单_*.xlsx")
        files = glob.glob(pattern)

        if not files:
            result['warnings'] = 1
            result['details'].append({
                'type': 'warning',
                'message': "未找到入库单文件"
            })
            return result

        # 核验每个入库单
        for file_path in files:
            result['total'] += 1
            try:
                details = self._validate_one_inbound_form(file_path)
                if details['status'] == 'passed':
                    result['passed'] += 1
                elif details['status'] == 'failed':
                    result['failed'] += 1
                result['details'].append(details)
            except Exception as e:
                result['failed'] += 1
                result['details'].append({
                    'type': 'failed',
                    'file': file_path,
                    'message': f"核验失败: {str(e)}"
                })

        print(f"    入库单核验完成: 总数 {result['total']}, 通过 {result['passed']}, 失败 {result['failed']}")
        return result

    def _validate_one_inbound_form(self, file_path):
        """核验单个入库单"""
        wb = load_workbook(file_path)
        ws = wb.active

        # 从表头获取入库单号
        inbound_no = ws['D4'].value
        if not inbound_no:
            return {
                'type': 'failed',
                'file': file_path,
                'message': f"无法读取入库单号"
            }

        # 从数据明细表查找对应物料
        matched_rows = self.df[self.df.iloc[:, 23].astype(str).str.contains(str(inbound_no), na=False)]

        if len(matched_rows) == 0:
            return {
                'type': 'failed',
                'file': file_path,
                'message': f"数据明细表中未找到入库单号 {inbound_no}"
            }

        errors = []
        # 检查物料数量是否匹配
        expected_count = len(matched_rows)

        # 计算实际物料数量（从第8行开始，读取非空行）
        actual_count = 0
        row_idx = 8
        while row_idx <= 100:  # 最多检查100行
            material_code = ws[f'B{row_idx}'].value
            if material_code and str(material_code).strip():
                actual_count += 1
            else:
                break
            row_idx += 1

        if expected_count != actual_count:
            errors.append(f"物料数量不匹配: 预期{expected_count}个, 实际{actual_count}个")

        if errors:
            return {
                'type': 'failed',
                'file': file_path,
                'inbound_no': inbound_no,
                'errors': errors
            }

        return {
            'type': 'passed',
            'file': file_path,
            'inbound_no': inbound_no,
            'material_count': actual_count
        }

    def _validate_outbound_forms(self):
        """核验出库单"""
        print("\n  核验出库单...")

        result = {
            'total': 0,
            'passed': 0,
            'failed': 0,
            'warnings': 0,
            'details': []
        }

        # 查找所有出库单文件
        pattern = os.path.join(self.output_dir, "出库单_*.xlsx")
        files = glob.glob(pattern)

        if not files:
            result['warnings'] = 1
            result['details'].append({
                'type': 'warning',
                'message': "未找到出库单文件"
            })
            return result

        # 核验每个出库单
        for file_path in files:
            result['total'] += 1
            try:
                details = self._validate_one_outbound_form(file_path)
                if details['status'] == 'passed':
                    result['passed'] += 1
                elif details['status'] == 'failed':
                    result['failed'] += 1
                result['details'].append(details)
            except Exception as e:
                result['failed'] += 1
                result['details'].append({
                    'type': 'failed',
                    'file': file_path,
                    'message': f"核验失败: {str(e)}"
                })

        print(f"    出库单核验完成: 总数 {result['total']}, 通过 {result['passed']}, 失败 {result['failed']}")
        return result

    def _validate_one_outbound_form(self, file_path):
        """核验单个出库单"""
        wb = load_workbook(file_path)
        ws = wb.active

        # 从表头获取出库单号
        outbound_no = ws['D3'].value
        if not outbound_no:
            return {
                'type': 'failed',
                'file': file_path,
                'message': f"无法读取出库单号"
            }

        # 从数据明细表查找对应物料
        matched_rows = []
        for idx, row in self.df.iterrows():
            # 检查三个出库批次
            for col_idx in [25, 29, 33]:  # Z列、AD列、AH列
                out_no = self._get_value(row, col_idx)
                if out_no and str(outbound_no) in str(out_no):
                    matched_rows.append((idx, row))
                    break

        if len(matched_rows) == 0:
            return {
                'type': 'failed',
                'file': file_path,
                'message': f"数据明细表中未找到出库单号 {outbound_no}"
            }

        errors = []
        # 检查物料数量是否匹配
        expected_count = len(matched_rows)

        # 计算实际物料数量
        actual_count = 0
        row_idx = 6
        while row_idx <= 100:
            material_code = ws[f'B{row_idx}'].value
            if material_code and str(material_code).strip():
                actual_count += 1
            else:
                break
            row_idx += 1

        if expected_count != actual_count:
            errors.append(f"物料数量不匹配: 预期{expected_count}个, 实际{actual_count}个")

        if errors:
            return {
                'type': 'failed',
                'file': file_path,
                'outbound_no': outbound_no,
                'errors': errors
            }

        return {
            'type': 'passed',
            'file': file_path,
            'outbound_no': outbound_no,
            'material_count': actual_count
        }

    def _get_value(self, row, col_idx):
        """获取单元格值"""
        if col_idx >= len(row):
            return None
        val = row.iloc[col_idx]
        if pd.isna(val):
            return None
        return val
