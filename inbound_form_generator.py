"""
入库单生成器
相同入库单号、入库日期、仓库类型的物料可以合并到一张入库单
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import os


class InboundFormGenerator:
    """入库单生成器"""

    def __init__(self, data_file: str, output_dir: str, template_file: str = None):
        """
        初始化生成器

        Args:
            data_file: 数据明细表文件路径
            output_dir: 输出目录
            template_file: 入库单模板文件路径
        """
        self.data_file = data_file
        self.output_dir = output_dir
        self.template_file = template_file or "assets/入库单单模板.xlsx"

        # 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)

        # 加载模板
        self.template_wb = load_workbook(self.template_file)
        self.template_ws = self.template_wb.active

        # 加载数据
        self.df = pd.read_excel(data_file, header=3)
        self.df = self.df.dropna(how='all')

        # 建立列名到列索引的映射
        self.column_map = {}
        for idx, col_name in enumerate(self.df.columns):
            self.column_map[str(col_name).strip()] = idx

    def _get_column_index(self, possible_names, fallback_idx):
        """根据可能的列名查找列索引"""
        for name in possible_names:
            for col_name in self.df.columns:
                if name in str(col_name):
                    return self.df.columns.get_loc(col_name)
        return fallback_idx if fallback_idx < len(self.df.columns) else 0

    def generate_all(self):
        """
        生成所有入库单（按入库单号、入库日期、仓库类型分组）

        Returns:
            生成结果字典
        """
        print(f"开始生成入库单，共 {len(self.df)} 个物料")

        result = {
            'total': 0,
            'materials_count': len(self.df),
            'success': 0,
            'failed': 0,
            'forms': [],
            'errors': []
        }

        # 按入库单号、入库日期、仓库类型分组
        groups = self._group_materials()

        print(f"共生成 {len(groups)} 张入库单")

        for group_key, group_data in groups.items():
            try:
                file_path = self._generate_one(group_data, group_key)
                result['total'] += 1
                result['success'] += 1
                result['forms'].append({
                    'inbound_no': group_key[0],
                    'inbound_date': group_key[1],
                    'warehouse': group_key[2],
                    'material_count': len(group_data),
                    'file_path': file_path
                })
                print(f"  ✓ 生成成功: 入库单号 {group_key[0]}, 仓库 {group_key[2]}, 物料数 {len(group_data)}")
            except Exception as e:
                result['failed'] += 1
                error_msg = f"入库单 {group_key[0]} 生成失败: {str(e)}"
                result['errors'].append(error_msg)
                print(f"  ✗ {error_msg}")

        print(f"\n入库单生成完成！成功: {result['success']}, 失败: {result['failed']}")
        return result

    def _group_materials(self):
        """
        按入库单号、入库日期、仓库类型分组

        Returns:
            分组字典 {(入库单号, 入库日期, 仓库类型): [数据行列表]}
        """
        groups = {}

        # 列索引（根据数据明细表测试.xlsx的列结构）
        inbound_no_col = 23  # 入库单号
        inbound_date_col = 17  # 入库日期
        warehouse_col = 22  # 仓库

        print(f"    使用列索引: 入库单号={inbound_no_col}, 入库日期={inbound_date_col}, 仓库={warehouse_col}")

        for idx, row in self.df.iterrows():
            # 获取分组键
            inbound_no = self._get_value(row, inbound_no_col)
            inbound_date = self._get_value(row, inbound_date_col)
            warehouse = self._get_value(row, warehouse_col)

            # 格式化日期
            if inbound_date:
                inbound_date = self._format_date(inbound_date)

            # 跳过没有入库单号的物料
            if not inbound_no:
                continue

            group_key = (inbound_no, inbound_date, warehouse)

            if group_key not in groups:
                groups[group_key] = []
            groups[group_key].append(row)

        return groups

    def _generate_one(self, group_data, group_key):
        """
        生成一张入库单

        Args:
            group_data: 分组数据列表
            group_key: 分组键 (入库单号, 入库日期, 仓库类型)

        Returns:
            生成的文件路径
        """
        # 复制模板
        wb = load_workbook(self.template_file)
        ws = wb.active

        inbound_no, inbound_date, warehouse = group_key

        # 列索引（根据数据明细表测试.xlsx的列结构）
        supplier_col = 6  # 供应商
        purchase_order_col = 16  # 采购订单
        material_code_col = 1  # 物料编码
        material_name_col = 2  # 物料名称
        brand_col = 5  # 品牌
        spec_col = 3  # 规格
        unit_col = 4  # 单位
        qty_col = 18  # 入库数量
        batch_col = 19  # 入库批号
        supplier_batch_col = 20  # 供应商批次
        expiry_col = 21  # 有效期

        # 获取第一行的基本信息（同一组的这些信息应该相同）
        first_row = group_data[0]
        supplier = self._get_value(first_row, supplier_col)
        purchase_order = self._get_value(first_row, purchase_order_col)

        # 填充表头信息（根据模板文件的结构）
        # 行4: 公司名称(C4)、入库类型(E4)、入库单号(H4)
        ws['H4'] = inbound_no

        # 行5: 采购订单(C5)、仓库(E5)、入库日期(H5)
        ws['C5'] = purchase_order
        ws['E5'] = warehouse
        ws['H5'] = inbound_date

        # 行6: 供应商(C6)、备注(E6)
        ws['C6'] = supplier
        ws['E6'] = ''

        # 填充物料明细（从第8行开始）
        start_row = 8
        for i, row in enumerate(group_data):
            row_idx = start_row + i

            # 获取物料数据
            seq = (i + 1) * 10  # 序号：10, 20, 30...
            material_code = self._get_value(row, material_code_col)
            material_name = self._get_value(row, material_name_col)
            brand = self._get_value(row, brand_col)
            specification = self._get_value(row, spec_col)
            unit = self._get_value(row, unit_col)
            quantity = self._get_value(row, qty_col)
            batch_no = self._get_value(row, batch_col)
            supplier_batch = self._get_value(row, supplier_batch_col)
            expiry_date = self._get_value(row, expiry_col)

            # 填充到对应的列
            ws.cell(row=row_idx, column=1, value=seq)  # 序号
            ws.cell(row=row_idx, column=2, value=material_code)  # 物料编码
            ws.cell(row=row_idx, column=3, value=material_name)  # 物料名称
            ws.cell(row=row_idx, column=4, value=brand)  # 品牌
            ws.cell(row=row_idx, column=5, value=specification)  # 规格
            ws.cell(row=row_idx, column=6, value=unit)  # 单位
            ws.cell(row=row_idx, column=7, value=self._format_number(quantity))  # 数量
            ws.cell(row=row_idx, column=8, value=batch_no)  # 批号
            ws.cell(row=row_idx, column=9, value=supplier_batch)  # 供应商批号
            ws.cell(row=row_idx, column=10, value=self._format_date(expiry_date))  # 失效日期

        # 生成文件名
        filename = f"入库单_{inbound_no}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(self.output_dir, filename)

        # 保存文件
        wb.save(file_path)

        return file_path



    def _get_value(self, row, col_idx):
        """获取单元格值"""
        if col_idx >= len(row):
            return None
        val = row.iloc[col_idx]
        if pd.isna(val):
            return None
        return val

    def _format_date(self, date_value):
        """格式化日期"""
        if pd.isna(date_value):
            return ''
        if isinstance(date_value, datetime):
            return date_value.strftime('%Y-%m-%d')
        if isinstance(date_value, str):
            # 尝试解析
            for fmt in ['%Y.%m.%d', '%Y-%m-%d', '%Y/%m/%d']:
                try:
                    dt = datetime.strptime(date_value, fmt)
                    return dt.strftime('%Y-%m-%d')
                except:
                    continue
        return str(date_value)

    def _format_number(self, num_value):
        """格式化数字，保留两位小数"""
        if pd.isna(num_value):
            return ''
        if isinstance(num_value, (int, float)):
            return f"{float(num_value):.2f}"
        if isinstance(num_value, str) and num_value.strip():
            # 尝试转换
            try:
                return f"{float(num_value):.2f}"
            except:
                return num_value
        return ''
